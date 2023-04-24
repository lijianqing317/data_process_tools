import os
import json
import glob
import copy
import numpy as np
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl import Workbook, load_workbook
import openpyxl.drawing.image
import seaborn
import matplotlib.pyplot as plt
import traceback
import shutil


class MyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return super(MyEncoder, self).default(obj)


# 缺陷统计信息类
class ObjDefectReportInfo(object):
    def __init__(self, class_id, class_name):
        super(ObjDefectReportInfo, self).__init__()
        self.class_id = class_id
        self.class_name = class_name
        self.gt_num = 0
        self.total_detection_num = 0
        self.correct_detection_num = 0
        self.missed_detection_num = 0  # 漏检
        self.missed_detection_rate = 0.0
        self.confused_detection_dict = {}  # 错检
        self.over_detection_num = 0  # 过检
        self.model_over_detection_rate = 0.0
        self.field_over_detection_num = 0
        self.field_over_detection_rate = 0.0


def result_to_save_json(image_name, bboxes, class_name, json_base_path, img_bash_path):
    """
    @param image_name: image_name   must
    @param image: image data        must
    @param result: model result     must
    @param json_base_path: model to json must
    """
    # image h * w
    # image_h, image_w = image.shape[0], image.shape[1]
    image_h = 500
    image_w = 500
    dic_all = {}
    shapes = []
    shapes_over = []
    shapes_miss = []
    shapes_error = []
    for idx, bbox in enumerate(bboxes):

        shape = {}
        points = []

        if (len(bbox) > 4):
            shape = bbox[-1]
            # print(shape)
            shape["label"] = class_name[idx]
            shapes.append(shape)
            if "miss" in class_name[idx]:
                shapes_miss.append(shape)
            else:
                shapes_error.append(shape)
        else:
            x1, y1, x2, y2 = bbox[0], bbox[1], bbox[2], bbox[3]
            shape["width"] = 1
            shape["shape_type"] = "rectangle"
            shape["group_id"] = ""
            shape["label"] = class_name[idx]
            # shape["score"] = round(score, 2)

            points.append([x1, y1])
            points.append([x2, y2])
            shape["points"] = points
            shapes.append(shape)
            shapes_over.append(shape)

    dic_all["version"] = "1.0"
    dic_all["flags"] = {}
    dic_all["shapes"] = shapes
    dic_all["imagePath"] = image_name
    dic_all["imageData"] = None
    dic_all["imageHeight"] = image_h
    dic_all["imageWidth"] = image_w
    dic_all["imageDepth"] = 1
    dic_all["imageLabeled"] = "true"

    json_name = image_name.split(".")[0] + ".json"
    json_path = os.path.join(json_base_path, "all", json_name)
    json_miss_path = os.path.join(json_base_path, "miss", json_name)
    json_error_path = os.path.join(json_base_path, "error", json_name)
    json_over_path = os.path.join(json_base_path, "over", json_name)
    if not os.path.exists(os.path.join(json_base_path, "all")):
        os.mkdir(os.path.join(json_base_path, "all"))
    if not os.path.exists(os.path.join(json_base_path, "miss")):
        os.mkdir(os.path.join(json_base_path, "miss"))
    if not os.path.exists(os.path.join(json_base_path, "error")):
        os.mkdir(os.path.join(json_base_path, "error"))
    if not os.path.exists(os.path.join(json_base_path, "over")):
        os.mkdir(os.path.join(json_base_path, "over"))

    img_old_path = os.path.join(img_bash_path, image_name)
    img_new_path = os.path.join(json_base_path, image_name)
    shutil.copy(img_old_path, img_new_path)
    # print(args.output_dir)
    with open(json_path, "w", encoding="utf-8") as f:
        # print(dic_all)
        content = json.dumps(dic_all, cls=MyEncoder, ensure_ascii=False, indent=4)
        f.write(content)
    if len(shapes_miss):
        dic_all["shapes"] = shapes_miss
        with open(json_miss_path, "w", encoding="utf-8") as f:
            content = json.dumps(dic_all, cls=MyEncoder, ensure_ascii=False, indent=4)
            f.write(content)
        img_new_miss_path = os.path.join(json_base_path, "miss", image_name)
        shutil.copy(img_old_path, img_new_miss_path)
    if len(shapes_over):
        dic_all["shapes"] = shapes_over
        with open(json_over_path, "w", encoding="utf-8") as f:
            content = json.dumps(dic_all, cls=MyEncoder, ensure_ascii=False, indent=4)
            f.write(content)
        img_new_over_path = os.path.join(json_base_path, "over", image_name)
        shutil.copy(img_old_path, img_new_over_path)
    if len(shapes_error):
        dic_all["shapes"] = shapes_error
        with open(json_error_path, "w", encoding="utf-8") as f:
            content = json.dump(dic_all, f, cls=MyEncoder, ensure_ascii=False, indent=4)
        img_new_error_path = os.path.join(json_base_path, "error", image_name)
        shutil.copy(img_old_path, img_new_error_path)


# 全局类别id和类别名字典
class_name_id_dict = {}
class_id_name_dict = {}


def convert_class_id_name_dict():
    global class_id_name_dict
    class_id_name_dict = {v: k for k, v in class_name_id_dict.items()}


def parse_bbox_in_json_shape(shape, shape_type=None, w=65535, h=65535):
    """get label bbox and class name from label shape info.
        Args:
            shape (:dict): The defect label shape info dict
            shape_type (str): The defect label shape type.
            w (int): the image width.
            h (int): the image height.
        Returns:
            class_name, x1, y1, x2, y2
        """
    class_name = shape["label"]
    shape_type = shape["shape_type"] if shape_type is None else shape_type
    points = shape["points"]
    x_min = 0
    y_min = 0
    x_max = w
    y_max = h
    if shape_type == "circle":  # 圆形标注需要特殊处理
        tmp = (
                      (points[1][1] - points[0][1]) ** 2 + (points[1][0] - points[0][0]) ** 2
              ) ** 0.5
        x1 = max(points[0][0] - tmp, x_min)
        x2 = min(points[0][0] + tmp, x_max)
        y1 = max(points[0][1] - tmp, y_min)
        y2 = min(points[0][1] + tmp, y_max)
    elif shape_type in ["rectangle", "polygon", "linestrip", "line", "point"]:
        # 一般标注类型只取左上角和右下角作为gt标注矩形框
        x1, y1, x2, y2 = 65535, 65535, -65535, -65535
        for point in points:
            x1, y1, x2, y2 = (
                max(min(x1, point[0]), x_min),
                max(min(y1, point[1]), y_min),
                min(max(x2, point[0]), x_max),
                min(max(y2, point[1]), y_max),
            )
        if x1 == x2:
            x1 = max(x1 - 2, x_min)
            x2 = min(x1 + 2, x_max)
        if y1 == y2:
            y1 = max(y1 - 2, y_min)
            y2 = min(y2 + 2, y_max)

    else:
        raise Exception(f"Unknown shape type: {shape_type}")
    return class_name, x1, y1, x2, y2


# 解析gt的标注信息
def parse_gt_label_info(gt_label_dict):
    if "shapes" not in gt_label_dict or len(gt_label_dict["shapes"]) == 0:
        return None
    gt_labels = []

    shapes = gt_label_dict["shapes"]
    # img_w = gt_label_dict["imageWidth"]
    # img_h = gt_label_dict["imageHeight"]
    # img_w = 65535
    # img_h = 65535
    for shape in shapes:
        class_name, x1, y1, x2, y2 = parse_bbox_in_json_shape(shape)
        if class_name in class_name_id_dict:
            class_id = class_name_id_dict[class_name]
        else:
            class_id = len(class_name_id_dict)
            class_name_id_dict[class_name] = class_id
        gt_labels.append([class_id, x1, y1, x2, y2, shape])
    return gt_labels


# 解析模型输出的预测信息
def parse_prediction_info(val_prediction_dict):
    if "shapes" not in val_prediction_dict or len(val_prediction_dict["shapes"]) == 0:
        return None

    detections = []

    shapes = val_prediction_dict["shapes"]
    for shape in shapes:
        class_name, x1, y1, x2, y2 = parse_bbox_in_json_shape(shape, "rectangle")
        if class_name in class_name_id_dict:
            class_id = class_name_id_dict[class_name]
        else:
            class_id = len(class_name_id_dict)
            class_name_id_dict[class_name] = class_id
        conf = shape["score"]
        detections.append([x1, y1, x2, y2, conf, class_id])
    return detections


# iou计算
def compute_iou(bbox1, bbox2):
    s_rec1 = (bbox1[2] - bbox1[0]) * (bbox1[3] - bbox1[1])
    s_rec2 = (bbox2[2] - bbox2[0]) * (bbox2[3] - bbox2[1])

    x0 = max(bbox1[0], bbox2[0])
    y0 = max(bbox1[1], bbox2[1])
    x1 = min(bbox1[2], bbox2[2])
    y1 = min(bbox1[3], bbox2[3])

    w = max(0, x1 - x0)
    h = max(0, y1 - y0)
    iou = (w * h) / (s_rec1 + s_rec2 - w * h)
    return iou


# 计算gt和预测框之间的匹配关系
# 返回： gt和预测匹配上的索引字典  和  经过阈值被过滤掉的预测框索引列表
def get_matches_between_gt_and_label(
        gt_labels, detections, confidence=0.5, iou_threshold=0.5
):
    if gt_labels is None or detections is None:
        return {}, []
    gt_labels = copy.deepcopy(gt_labels)
    detections = copy.deepcopy(detections)

    # 定义被gt已经匹配上的检测框索引集合
    selected_detection_idx = set()
    # 定义gt和预测匹配上的索引字典
    gt_det_matched_idx_dict = {}
    # 定义需要被过滤的预测框索引集合
    discarded_detection_idx = set()
    for gt_idx, gt_label in enumerate(gt_labels):  # 遍历gt框
        best_detection_idx = 0
        max_iou = 0.0
        for detection_idx, detection in enumerate(detections):  # 遍历预测框
            # if detection_idx in selected_detection_idx:  # 如果预测框已经被gt匹配过了 就不需要再匹配了（这个逻辑后面可以改）
            #     continue
            if detection[4] < confidence:  # 如果预测框分值低于置信度，则舍弃预测框
                discarded_detection_idx.add(detection_idx)
                continue
            iou = compute_iou(gt_label[1:], detection[:4])  # 计算预测框和gt的iou值
            if iou == 0.0:  # 如果iou=0没有交集，则算过检，不需要丢弃预测框
                continue
            elif iou < iou_threshold:  # 如果低于iou阈值，则舍弃预测框
                discarded_detection_idx.add(detection_idx)
                continue
            if iou > max_iou:  # 选取iou最高的预测框和gt匹配
                max_iou = iou
                best_detection_idx = detection_idx
        if max_iou >= iou_threshold:
            selected_detection_idx.add(best_detection_idx)
            gt_det_matched_idx_dict[gt_idx] = best_detection_idx
            if best_detection_idx in discarded_detection_idx:  # 如果当前预测框在之前被视为丢弃框，但又和当前gt框匹配上，则需要将该预测框不再视为丢弃框
                discarded_detection_idx.remove(best_detection_idx)

    return gt_det_matched_idx_dict, list(discarded_detection_idx)


# 统计每张图的各项指标信息
def parse_image_detection_report(gt_labels, detections, matched_idx_dict, discarded_detection_idx=None):
    result_bbox = []
    result_class_name = []
    matched_gt_idx = matched_idx_dict.keys()  # 匹配上的gt框索引列表
    matched_detection_idx = matched_idx_dict.values()  # 匹配上的预测框索引列表
    discarded_detection_idx = [] if (discarded_detection_idx is None or len(discarded_detection_idx) == 0) \
        else discarded_detection_idx

    image_class_defect_infos = {}  # 定义不同缺陷类别的统计指标信息
    for gt_idx, detection_idx in matched_idx_dict.items():
        gt_cls_id = gt_labels[gt_idx][0]
        gt_cls_name = class_id_name_dict[gt_cls_id]
        detection_cls_id = detections[detection_idx][-1]
        detection_cls_name = class_id_name_dict[detection_cls_id]

        if gt_cls_id == detection_cls_id:  # 如果gt类别和预测类别一致
            cls_count_info = (
                image_class_defect_infos[gt_cls_id]
                if gt_cls_id in image_class_defect_infos
                else ObjDefectReportInfo(gt_cls_id, gt_cls_name)
            )
            cls_count_info.gt_num += 1
            cls_count_info.total_detection_num += 1
            cls_count_info.correct_detection_num += 1
            image_class_defect_infos[gt_cls_id] = cls_count_info
        else:  # 如果gt类别和预测类别不一致，需要分别统计信息，错检出
            gt_cls_count_info = (
                image_class_defect_infos[gt_cls_id]
                if gt_cls_id in image_class_defect_infos
                else ObjDefectReportInfo(gt_cls_id, gt_cls_name)
            )
            detection_cls_count_info = (
                image_class_defect_infos[detection_cls_id]
                if detection_cls_id in image_class_defect_infos
                else ObjDefectReportInfo(detection_cls_id, detection_cls_name)
            )
            result_bbox.append(detections[detection_idx][:4] + [gt_labels[gt_idx][-1]])
            result_class_name.append("error_" + detection_cls_name + "@" + gt_cls_name)
            gt_cls_count_info.gt_num += 1
            detection_cls_count_info.total_detection_num += 1
            # 统计混淆类别数目
            if gt_cls_id in detection_cls_count_info.confused_detection_dict:
                detection_cls_count_info.confused_detection_dict[gt_cls_id] += 1
            else:
                detection_cls_count_info.confused_detection_dict[gt_cls_id] = 1
            image_class_defect_infos[gt_cls_id] = gt_cls_count_info
            image_class_defect_infos[detection_cls_id] = detection_cls_count_info

    # 未匹配上的gt索引列表
    miss_matched_gt_idx_list = (
        [idx for idx, item in enumerate(gt_labels) if idx not in matched_gt_idx]
        if gt_labels
        else []
    )

    # 未匹配上的预测框索引列表
    miss_matched_detection_idx_list = (
        [idx for idx, item in enumerate(detections) if
         (idx not in matched_detection_idx and idx not in discarded_detection_idx)]
        if detections
        else []
    )
    # gt没匹配，漏检
    for miss_matched_gt_idx in miss_matched_gt_idx_list:
        gt_item = gt_labels[miss_matched_gt_idx]
        gt_cls_id = gt_item[0]
        gt_cls_name = class_id_name_dict[gt_cls_id]
        cls_count_info = (
            image_class_defect_infos[gt_cls_id]
            if gt_cls_id in image_class_defect_infos
            else ObjDefectReportInfo(gt_cls_id, gt_cls_name)
        )
        cls_count_info.gt_num += 1
        cls_count_info.missed_detection_num += 1
        image_class_defect_infos[gt_cls_id] = cls_count_info
        result_bbox.append(gt_item[1:])
        result_class_name.append("miss@" + gt_cls_name)
    # 过检
    for miss_matched_detection_idx in miss_matched_detection_idx_list:
        miss_det_item = detections[miss_matched_detection_idx]
        miss_det_cls_id = miss_det_item[5]
        miss_det_cls_name = class_id_name_dict[miss_det_cls_id]
        cls_count_info = (
            image_class_defect_infos[miss_det_cls_id]
            if miss_det_cls_id in image_class_defect_infos
            else ObjDefectReportInfo(miss_matched_detection_idx, miss_det_cls_name)
        )
        cls_count_info.total_detection_num += 1
        cls_count_info.over_detection_num += 1
        image_class_defect_infos[miss_det_cls_id] = cls_count_info
        result_bbox.append(miss_det_item[:4])
        result_class_name.append("overkill")
        #result_class_name.append("over@" + miss_det_cls_name)
    return image_class_defect_infos, result_bbox, result_class_name


# 基于统计好的每张图的指标信息，统计所有类别的指标信息
def statistics_all_defect_class_infos(defect_img_infos):
    defect_classes_infos = {}
    for image_name, image_detection_info in defect_img_infos.items():
        for class_id, class_count_info in image_detection_info.items():
            class_name = class_id_name_dict[class_id]
            defect_class_info = (
                defect_classes_infos[class_id]
                if class_id in defect_classes_infos
                else ObjDefectReportInfo(class_id, class_name)
            )

            defect_class_info.gt_num += class_count_info.gt_num  # 类别总gt数目
            defect_class_info.total_detection_num += (  # 类别总预测数目
                class_count_info.total_detection_num
            )
            defect_class_info.correct_detection_num += (  # 类别总预测正确数目
                class_count_info.correct_detection_num
            )
            defect_class_info.missed_detection_num += (  # 类别总漏检数目
                class_count_info.missed_detection_num
            )
            defect_class_info.over_detection_num += class_count_info.over_detection_num  # 类别总过检数目
            for (
                    confused_cls_id,
                    confused_num,
            ) in class_count_info.confused_detection_dict.items():
                defect_class_confused_num = (  # 类别总检测混淆数目
                    defect_class_info.confused_detection_dict[confused_cls_id]
                    if confused_cls_id in defect_class_info.confused_detection_dict
                    else 0
                )
                defect_class_confused_num += confused_num
                defect_class_info.confused_detection_dict[
                    confused_cls_id
                ] = defect_class_confused_num

            defect_class_info.missed_detection_rate = (
                defect_class_info.missed_detection_num / defect_class_info.gt_num
                if defect_class_info.gt_num != 0
                else 0
            )
            defect_class_info.model_over_detection_rate = (
                defect_class_info.over_detection_num
                / defect_class_info.total_detection_num
                if defect_class_info.total_detection_num != 0
                else 0
            )
            # if defect_class_info.over_detection_num > 0:
            #     defect_class_info.field_over_detection_num += 1

            defect_classes_infos[class_id] = defect_class_info
    return defect_classes_infos


# 生成excel文件
def save_report_as_excel_file(defect_img_infos, save_path):
    header_font = Font(
        name="微软雅黑",
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline=None,
        strike=False,
        color="FF000000",
    )
    border = Border(
        # left=Side(style='medium', color='FF000000'), right=Side(style='medium', color='FF000000'),
        top=Side(style="medium", color="5C799D"),
        bottom=Side(style="medium", color="5C799D"),
        diagonal=Side(style="medium", color="FF000000"),
        diagonal_direction=0,
        outline=Side(style="medium", color="FF000000"),
        vertical=Side(style="medium", color="FF000000"),
        horizontal=Side(style="medium", color="FF000000"),
    )
    alignment = Alignment(horizontal="center", vertical="center")

    num_str_dict = {}
    a_z = [chr(a) for a in range(ord("A"), ord("Z") + 1)]
    aa_az = ["A" + chr(a) for a in range(ord("A"), ord("Z") + 1)]
    a_az = a_z + aa_az
    for i in a_az:
        num_str_dict[a_az.index(i) + 1] = i

    wb = Workbook()
    ws = wb.active

    class_id_list = defect_img_infos.keys()
    class_name_list = [class_id_name_dict[i] for i in class_id_list]
    ws.append(["","all"] + class_name_list)

    col_header = ["", "GT数量", "漏检数", "漏检率", "检出数", "过检数", "模型过检率"]
    for i, n in enumerate(col_header):
        ws.cell(i + 1, 1).value = n
    sum_gts=0
    sum_miss=0
    sum_dets=0
    sum_overkill=0
    for i, class_id in enumerate(class_id_list):
        class_defect_info = defect_img_infos[class_id]
        ws.cell(2, i + 3).value = class_defect_info.gt_num
        ws.cell(3, i + 3).value = class_defect_info.missed_detection_num
        ws.cell(4, i + 3).value = f'{round(class_defect_info.missed_detection_rate * 100, 2)}%'
        ws.cell(5, i + 3).value = class_defect_info.total_detection_num
        ws.cell(6, i + 3).value = class_defect_info.over_detection_num
        ws.cell(7, i + 3).value = f'{round(class_defect_info.model_over_detection_rate * 100, 2)}%'
        sum_gts +=class_defect_info.gt_num
        sum_miss +=class_defect_info.missed_detection_num
        sum_dets +=class_defect_info.total_detection_num
        sum_overkill +=class_defect_info.over_detection_num
    ws.cell(2, 2).value = sum_gts
    ws.cell(3, 2).value = sum_miss
    ws.cell(4, 2).value = f'{round(sum_miss/sum_gts * 100, 2)}%'
    ws.cell(5, 2).value = sum_dets
    ws.cell(6, 2).value = sum_overkill
    ws.cell(7, 2).value = f'{round(sum_overkill/sum_dets * 100, 2)}%'

    num_cols = ws.max_column
    num_rows = ws.max_row
    for c in range(1, num_cols + 1):
        for r in range(1, num_rows + 1):
            if r == 1 or c == 1:
                ws.cell(r, c).font = header_font
                if r == 1:
                    ws.column_dimensions[num_str_dict[c]].width = 20
                    ws.cell(r, c).fill = PatternFill("solid", fgColor="218FC5")
                if c == 1:
                    ws.row_dimensions[r].height = 25
            ws.cell(r, c).border = border
            if r != 1 and c != 1:
                ws.cell(r, c).font = Font(
                    name="宋体",
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline="none",
                    strike=False,
                    color="FF000000",
                )
            ws.cell(r, c).alignment = alignment

    wb.save(save_path)

    return [sum_gts, sum_miss, f'{round(sum_miss / sum_gts * 100, 2)}%', sum_dets, sum_overkill,
            f'{round(sum_overkill / sum_dets * 100, 2)}%']

#生成整体阈值文档
def save_report_as_excel_file_all(all_dets_dic, save_path):
    header_font = Font(
        name="微软雅黑",
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline=None,
        strike=False,
        color="FF000000",
    )
    border = Border(
        # left=Side(style='medium', color='FF000000'), right=Side(style='medium', color='FF000000'),
        top=Side(style="medium", color="5C799D"),
        bottom=Side(style="medium", color="5C799D"),
        diagonal=Side(style="medium", color="FF000000"),
        diagonal_direction=0,
        outline=Side(style="medium", color="FF000000"),
        vertical=Side(style="medium", color="FF000000"),
        horizontal=Side(style="medium", color="FF000000"),
    )
    alignment = Alignment(horizontal="center", vertical="center")

    num_str_dict = {}
    a_z = [chr(a) for a in range(ord("A"), ord("Z") + 1)]
    aa_az = ["A" + chr(a) for a in range(ord("A"), ord("Z") + 1)]
    a_az = a_z + aa_az
    for i in a_az:
        num_str_dict[a_az.index(i) + 1] = i

    wb = Workbook()
    ws = wb.active

    class_id_list = all_dets_dic.keys()
    print(class_id_list,'===============')
    print(all_dets_dic,'======1=========')
    class_name_list = [str(i) for i in class_id_list]
    ws.append(["",] + class_name_list)
    print(class_id_list, '=======2========')
    col_header = ["", "GT数量", "漏检数", "漏检率", "检出数", "过检数", "模型过检率"]
    for i, n in enumerate(col_header):
        ws.cell(i + 1, 1).value = n
    for i, class_id in enumerate(class_id_list):
        class_defect_info = all_dets_dic[class_id]
        for j ,m in  enumerate(class_defect_info):
            ws.cell(j+2, i + 2).value = m

    num_cols = ws.max_column
    num_rows = ws.max_row
    for c in range(1, num_cols + 1):
        for r in range(1, num_rows + 1):
            if r == 1 or c == 1:
                ws.cell(r, c).font = header_font
                if r == 1:
                    ws.column_dimensions[num_str_dict[c]].width = 20
                    ws.cell(r, c).fill = PatternFill("solid", fgColor="218FC5")
                if c == 1:
                    ws.row_dimensions[r].height = 25
            ws.cell(r, c).border = border
            if r != 1 and c != 1:
                ws.cell(r, c).font = Font(
                    name="宋体",
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline="none",
                    strike=False,
                    color="FF000000",
                )
            ws.cell(r, c).alignment = alignment

    wb.save(save_path)


# 画混淆矩阵
def plot_matrix(matrix, class_name_list, insert_excel_file, show_flag=False):
    seaborn.set(font_scale=0.5)
    seaborn.heatmap(
        matrix,
        vmin=0,
        vmax=np.max(matrix),
        annot=True,
        linewidths=0.2,
        cmap="YlGnBu",
        fmt='.20g',
        cbar=True,
        xticklabels=class_name_list,
        yticklabels=class_name_list,
    )

    if show_flag:
        plt.show()
    temp_path = os.path.join(os.path.dirname(insert_excel_file), "temp_fig.jpg")
    plt.savefig(temp_path, dpi=150)

    wb = (
        load_workbook(insert_excel_file)
        if os.path.exists(insert_excel_file)
        else Workbook()
    )
    ws = wb.active

    ws.add_image(openpyxl.drawing.image.Image(temp_path), f"C{ws.max_row + 1}")
    wb.save(insert_excel_file)
    os.remove(temp_path)


# 生成混淆矩阵
def generate_confused_matrix(defect_classes_infos):
    detection_class_id_list = defect_classes_infos.keys()
    detection_class_id_list = [-1] + sorted(detection_class_id_list)  # 预测到的检测类别id列表，第一个元素是背景(-1)
    gt_class_id_list = [
        i for i in defect_classes_infos.keys() if defect_classes_infos[i].gt_num > 0
    ]
    gt_class_id_list = [-1] + sorted(gt_class_id_list)  # gt类别id列表，第一个元素是背景(-1)

    class_id_list = (
        detection_class_id_list
        if len(detection_class_id_list) > 0
        else gt_class_id_list
    )
    class_name_list = [class_id_name_dict[i] for i in class_id_list if i != -1]
    class_name_list = ["background"] + class_name_list
    matrix = np.zeros((len(class_id_list), len(class_id_list)))

    for i, gc in enumerate(class_id_list):
        for j, dc in enumerate(class_id_list):
            if gc == -1 and dc == -1:  # 第一行第一列，是背景对背景
                matrix[i, j] = 0
                continue
            elif gc == -1:  # 第一行背景，统计对应的过检数目
                matrix[i, j] = (
                    defect_classes_infos[dc].over_detection_num
                    if dc in defect_classes_infos
                    else 0
                )
            elif dc == -1:  # 第一列背景，统计对应的漏检数目
                matrix[i, j] = (
                    defect_classes_infos[gc].missed_detection_num
                    if gc in defect_classes_infos
                    else 0
                )
            else:
                if gc == dc:  # 列等于行时，统计识别正确数目
                    matrix[i, j] = (
                        defect_classes_infos[dc].correct_detection_num
                        if dc in defect_classes_infos
                        else 0
                    )
                else:
                    matrix[i, j] = (  # 列不等于行时，统计识别错误类别数目
                        (
                            defect_classes_infos[dc].confused_detection_dict[gc]
                            if gc in defect_classes_infos[dc].confused_detection_dict
                            else 0
                        )
                        if dc in defect_classes_infos
                        else 0
                    )

    return matrix, class_name_list


def matched_bbox_and_parse_val_infos(gt_labels, detections, defect_img_infos, img_base_name, confidence, iou_threshold,
                                     gt_dir, pr_path=""):
    convert_class_id_name_dict()  # 转换一下类别名-ID字典

    # 获取gt和预测框间的匹配关系，和需要丢弃的预测框
    matched_idx_dict, discarded_detection_idx = get_matches_between_gt_and_label(
        gt_labels,
        detections,
        confidence=confidence,
        iou_threshold=iou_threshold,
    )

    # 获取当前图片统计得到的每个缺陷类别指标数据
    image_class_defect_infos, json_bbox, json_label = parse_image_detection_report(
        gt_labels, detections, matched_idx_dict, discarded_detection_idx
    )
    if pr_path != "" and len(json_bbox) != 0:
        result_to_save_json(img_base_name, json_bbox, json_label, pr_path, gt_dir)
    # 将当前图片指标信息存入字典
    defect_img_infos[img_base_name] = image_class_defect_infos


def generate_evaluate_reports(
        gt_dir, val_dir, save_dir, pr_path="", iou_threshold=0.5, confidence=0.1
):
    defect_img_infos = {}
    if pr_path != "":
        if not os.path.exists(pr_path):
            os.mkdir(pr_path)

    gt_json_paths = glob.glob(f"{gt_dir}/*.json") if gt_dir is not None else []
    gt_json_names = [os.path.basename(i) for i in gt_json_paths]
    detection_json_paths = glob.glob(f"{val_dir}/*.json")
    missing_gt_json_paths = [i for i in detection_json_paths if os.path.basename(i) not in gt_json_names]

    if not os.path.exists(save_dir):
        os.mkdir(save_dir)

    for gt_label_path in gt_json_paths:
        try:
            gt_label_name = os.path.basename(gt_label_path)
            img_base_name = gt_label_name.replace(".json", ".jpg")

            with open(gt_label_path, "r") as f:
                gt_label_dict = json.load(f)
                gt_labels = parse_gt_label_info(gt_label_dict)  # 解析获得gt框

            val_prediction_path = os.path.join(val_dir, gt_label_name)
            if not os.path.exists(val_prediction_path):
                detections = None
            else:
                with open(val_prediction_path, "r") as f:
                    val_prediction_dict = json.load(f)
                    detections = parse_prediction_info(val_prediction_dict)  # 解析获得预测框

            # 匹配检测框 并解析图片各类别的检测结果
            matched_bbox_and_parse_val_infos(gt_labels, detections, defect_img_infos, img_base_name, confidence,
                                             iou_threshold, gt_dir, pr_path)
        except RuntimeError:
            print(
                f"ERROR! gt image path: {gt_label_path}, gt_labels: {gt_labels}, detections: {detections}\n "
                f"{traceback.format_exc()} "
            )

    # 对缺少对应gt_json的预测结果做处理
    for val_json_path in missing_gt_json_paths:
        with open(val_json_path, "r") as f:
            val_prediction_dict = json.load(f)
            over_detections = parse_prediction_info(val_prediction_dict)
            matched_bbox_and_parse_val_infos(None, over_detections, defect_img_infos,
                                             os.path.basename(val_json_path).replace(".json", ".jpg"),
                                             confidence, iou_threshold, gt_dir, pr_path)

    # 统计所有图片的缺陷类别指标信息
    all_class_defect_infos = statistics_all_defect_class_infos(defect_img_infos)
    # 生成混淆矩阵
    confused_matrix, class_name_list = generate_confused_matrix(all_class_defect_infos)
    # 生成excel文件报告
    all_det_list_per_sore=save_report_as_excel_file(
        all_class_defect_infos,
        os.path.join(
            save_dir, f"iou_{iou_threshold}_conf_{confidence}_detection_report.xlsx"
        ),
    )


    # 画混淆矩阵图
    plot_matrix(
        confused_matrix,
        class_name_list,
        os.path.join(
            save_dir, f"iou_{iou_threshold}_conf_{confidence}_detection_report.xlsx"
        ),
    )
    return all_det_list_per_sore
import argparse
def get_args():
    """
    This method is already deprecated.
    Returns:

    """
    parser = argparse.ArgumentParser(
        description="Deprecated: "
    )
    parser.add_argument("--origin-json-dir", default="/home/lijq/Desktop/data/O_ALL/workspace_single/data/train/Rlp", help="")
    parser.add_argument("--result-json-dir", default="/home/lijq/Desktop/data/O_ALL/workspace_single/data_pre/Rlp", help="")
    parser.add_argument("--report-dir", default="/home/lijq/Desktop/data/O_ALL/workspace_single/data_report/Olp_report", help="")
    parser.add_argument("--loushi-dir", default="/home/lijq/Desktop/data/O_ALL/workspace_single/data_report/Olp_loushi", help="")
    parser.add_argument("--iou-threshold", default=0.0001, help="")
    parser.add_argument("--score-threshold", default=[0.05], type=float,help="")
    parser.add_argument("--score-threshold-default", default=True, type=bool,help="")#使用该项时score-threshold失效
    parser.add_argument("--version-name", default='0311det', type=str,help="")#使用该项时score-threshold失效
    args = parser.parse_args()
    return args
# INPUT_IMG=/home/lijq/Desktop/data/byd/workspace/data_cropped/val/ipad2
# OUTPUT_DIR=/home/lijq/Desktop/data/byd/workspace/data_cropped/val/ipad2_again
if __name__ == "__main__":
    try:
        args=get_args()
        if args.score_threshold_default:
            all_score_thrs=[0.01,0.02,0.03,0.04, 0.05,0.06,0.07,0.08, 0.09,0.1,0.11,0.12,0.13,0.14,0.15, 0.16,0.17,0.18,0.19,0.2, 0.25, 0.27,0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95,]
            # all_score_thrs=[i*0.01 for i in range(len(args.score_threshold))]
        else:
            all_score_thrs=args.score_threshold

        all_dets_dic={}
        # for i in args.score_threshold:
        for i in all_score_thrs:
            report_dir=os.path.join(args.report_dir)
            loushi_dir=os.path.join(args.loushi_dir,str(i))
            if not os.path.exists(report_dir):
                os.makedirs(report_dir)

            try:
                if i!=0.1:
                    all_det_list_per_sore=generate_evaluate_reports(args.origin_json_dir,
                                              args.result_json_dir,
                                              report_dir,
                                              iou_threshold=float(args.iou_threshold), confidence=i)
                else:
                    if not os.path.exists(loushi_dir):
                        os.makedirs(loushi_dir)
                    all_det_list_per_sore = generate_evaluate_reports(args.origin_json_dir,
                                                                      args.result_json_dir,
                                                                      report_dir,
                                                                      loushi_dir,
                                                                      iou_threshold=float(args.iou_threshold),
                                                                      confidence=i)
                all_dets_dic[i]=all_det_list_per_sore
            except Exception as e:
                print('e1:',e)

        # 生成所有阈值的汇总表
        save_path = os.path.join(args.report_dir,'{}_all_thrs.xlsx'.format(args.version_name))
        save_report_as_excel_file_all(all_dets_dic, save_path)
    except Exception as e:
        print('e:', e)
